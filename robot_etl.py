import openpyxl
import os
import re
import pandas as pd
from datetime import datetime, timedelta, date
from openpyxl.styles import Font

# --- 1. CONFIGURACIÓN ---
RUTAS_BASE = [
    r"G:\.shortcut-targets-by-id\1H6coCC4GgCcvOGjxbvwnkh2A1xto5hcc\2026. Granjas\Registros_ Lotes Activos\Agroavicola Chi-Hen_ Registros\PRODUCCION CHIHEN",
    r"G:\.shortcut-targets-by-id\1H6coCC4GgCcvOGjxbvwnkh2A1xto5hcc\2026. Granjas\Registros_ Lotes Activos\Agropecuaria Nueva del Oriente_ Registros Lotes Activos\PRODUCCION AGNO",
    r"G:\.shortcut-targets-by-id\1H6coCC4GgCcvOGjxbvwnkh2A1xto5hcc\2026. Granjas\Registros_ Lotes Activos\Grupo Empresarial RRL_ Registros\PRODUCCIÓN RRL"
]

HOJA_PROD = "PROD-SEM"
HOJA_INI = "INF-INI"
FILA_SUPERIOR = 5  # <--- Fila de respaldo para títulos
FILA_TITULOS = 6   # <--- Fila principal de títulos
FILA_DATOS_INICIO = 8
LIMITE_COLUMNAS = 62

ayer = (datetime.now() - timedelta(days=1)).date()
ARCHIVO_SALIDA = r"C:\Users\Administrador\Desktop\HUPA DATA ANALYTICS\GRANJAS-WEB\DATA\Consolidado_Produccion_FINAL.xlsx"

def limpiar(t):
    return str(t).strip().lower() if t else ""

def ejecutar_robot_multicarpetas():
    if os.path.exists(ARCHIVO_SALIDA):
        try: os.remove(ARCHIVO_SALIDA)
        except: pass

    wb_dest = openpyxl.Workbook()
    ws_dest = wb_dest.active
    ws_dest.title = "Consolidado_Produccion"
    
    fila_destino = 2
    cabeceras_creadas = False

    for ruta in RUTAS_BASE:
        print(f"\n📂 REVISANDO CARPETA: {ruta}")
        if not os.path.exists(ruta): continue
            
        ruta_lower = ruta.lower()
        if "chihen" in ruta_lower: razon_social = "CHIHEN"
        elif "agno" in ruta_lower or "oriente" in ruta_lower: razon_social = "AGNO"
        elif "rrl" in ruta_lower: razon_social = "RRL"
        else: razon_social = "OTRO"

        archivos = [f for f in os.listdir(ruta) if f.endswith(".xlsx") and not f.startswith("~$")]

        for nombre_archivo in archivos:
            print(f"  -> Analizando: {nombre_archivo}")
            try:
                wb_ori = openpyxl.load_workbook(os.path.join(ruta, nombre_archivo), data_only=True)
                
                # --- EXTRAER DE INF-INI (PRODUCCIÓN) ---
                nombre_granja, ubicacion_granja, linea_aves = "N/A", "N/A", "N/A"
                if HOJA_INI in wb_ori.sheetnames:
                    ws_ini = wb_ori[HOJA_INI]
                    en_seccion_produccion = False
                    for fila_ini in list(ws_ini.values)[:50]:
                        f_limpia = [limpiar(c) for c in fila_ini]
                        if any("información en producción" in str(c) for c in f_limpia):
                            en_seccion_produccion = True
                        for c_idx, celda in enumerate(f_limpia):
                            if "línea de las aves" in celda:
                                linea_aves = str(fila_ini[c_idx+1]).strip() if c_idx+1 < len(fila_ini) else linea_aves
                            if en_seccion_produccion:
                                if "nombre de granja" in celda:
                                    nombre_granja = str(fila_ini[c_idx+1]).strip()
                                elif "ubicación granja" in celda:
                                    ubicacion_granja = str(fila_ini[c_idx+1]).strip()

                # --- EXTRAER DE PROD-SEM ---
                if HOJA_PROD in wb_ori.sheetnames:
                    ws_prod = wb_ori[HOJA_PROD]
                    filas_lista = list(ws_prod.values)
                    
                    if len(filas_lista) >= FILA_TITULOS:
                        match_nombre = re.search(r'\d+', nombre_archivo)
                        global_lote = int(match_nombre.group()) if match_nombre else 0

                        # --- LÓGICA DE TÍTULOS DE DOS NIVELES ---
                        f5 = filas_lista[FILA_SUPERIOR - 1] # Fila de arriba
                        f6 = filas_lista[FILA_TITULOS - 1]   # Fila de títulos
                        
                        titulos_finales = []
                        for i in range(LIMITE_COLUMNAS):
                            val_f6 = f6[i] if i < len(f6) else None
                            val_f5 = f5[i] if i < len(f5) else None
                            # SI FILA 6 ES NONE, USA FILA 5
                            nombre_col = val_f6 if val_f6 is not None else val_f5
                            if nombre_col is None: nombre_col = f"Columna_{i+1}"
                            titulos_finales.append(str(nombre_col).strip())

                        titulos_sucios = [limpiar(t) for t in titulos_finales]
                        idx_fin = next((i for i, t in enumerate(titulos_sucios) if "final" in t and "sem" in t), 1)
                        idx_saldo = next((i for i, t in enumerate(titulos_sucios) if "saldo" in t), 3)
                        idx_edad = next((i for i, t in enumerate(titulos_sucios) if "edad" in t), 2)

                        if not cabeceras_creadas:
                            cols_extra = ["ARCHIVO_ORIGEN", "RAZON_SOCIAL", "GRANJA", "UBICACION", "LINEA_AVES", "LOTE", "NUM_GALPON"]
                            for c_idx, nombre_c in enumerate(cols_extra, 1):
                                ws_dest.cell(1, c_idx, nombre_c).font = Font(bold=True)
                            for c_idx, t in enumerate(titulos_finales, 8):
                                ws_dest.cell(1, c_idx, t).font = Font(bold=True)
                            cabeceras_creadas = True

                        contador_tabla = 0 
                        for fila in filas_lista[FILA_DATOS_INICIO - 1:]:
                            fila_str = [limpiar(c) for c in fila]
                            if len(fila_str) > 0 and "inicio sem" in fila_str[0]:
                                contador_tabla += 1
                                continue

                            if len(fila) > max(idx_fin, idx_saldo, idx_edad):
                                f_fin = fila[idx_fin]
                                if isinstance(f_fin, datetime): f_fin = f_fin.date()
                                if isinstance(f_fin, date):
                                    try: 
                                        v_saldo = float(fila[idx_saldo]) if fila[idx_saldo] is not None else 0
                                        v_edad = int(fila[idx_edad]) if fila[idx_edad] is not None else 0
                                    except: continue

                                    if f_fin <= ayer and v_edad >= 19 and v_saldo > 0:
                                        ws_dest.cell(fila_destino, 1, nombre_archivo)
                                        ws_dest.cell(fila_destino, 2, razon_social)
                                        ws_dest.cell(fila_destino, 3, nombre_granja)
                                        ws_dest.cell(fila_destino, 4, ubicacion_granja)
                                        ws_dest.cell(fila_destino, 5, linea_aves)
                                        ws_dest.cell(fila_destino, 6, global_lote)
                                        ws_dest.cell(fila_destino, 7, global_lote if contador_tabla == 0 else contador_tabla)
                                        
                                        for col_idx in range(LIMITE_COLUMNAS):
                                            valor = fila[col_idx]
                                            celda = ws_dest.cell(fila_destino, 8 + col_idx, valor)
                                            if isinstance(valor, (datetime, date)):
                                                celda.number_format = 'DD/MM/YYYY'
                                        fila_destino += 1
                wb_ori.close()
            except Exception as e:
                print(f"⚠️ Error en {nombre_archivo}: {e}")

    wb_dest.save(ARCHIVO_SALIDA)
    print(f"\n✅ EXTRACCIÓN COMPLETADA SIN 'NONES'.")

if __name__ == "__main__":
    ejecutar_robot_multicarpetas()